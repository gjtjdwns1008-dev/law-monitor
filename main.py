import requests
import xml.etree.ElementTree as ET
import pandas as pd
from google import genai
from google.genai import types
import time
import os
import json
import re
from datetime import datetime, timedelta, timezone
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
import gspread
from oauth2client.service_account import ServiceAccountCredentials

# ==========================================
# 1. 환경 변수 및 설정
# ==========================================
LAW_API_KEY = os.environ.get("LAW_API_KEY")
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY")
WEBHOOK_URL = "https://hook.eu1.make.com/okarw4rcy9yusgxj44ogornxbdj8r51u"

# [V27 신규] 구글 시트 직접 제어용 환경 변수
GCP_SA_JSON = os.environ.get("GCP_SA_JSON")
GOOGLE_SHEET_ID = os.environ.get("GOOGLE_SHEET_ID")

# ==========================================
# 2. 날짜 및 신형 AI 세팅
# ==========================================
KST = timezone(timedelta(hours=9))
today = datetime.now(KST)
TARGET_DATE = today.strftime("%Y%m%d") 
SEARCH_DATE_RANGE = f"{TARGET_DATE}~{TARGET_DATE}" 
FILE_PREFIX = today.strftime("%Y년_%m월_%d일")

client = genai.Client(api_key=GEMINI_API_KEY)

# 👇👇👇 여기부터 통째로 복사해서 붙여넣기 (아무 띄어쓰기 없이 맨 왼쪽 줄에 딱 붙여서!) 👇👇👇
COLUMNS = ["시행일자", "소관부처", "법령명", "개정유형", "주요 제·개정내용", "법령 관련 국가기술자격 종목", 
           "활용도 분석 구분", "활용도 분석 상세", "근거 조문", "AI 신뢰도", "검토 필요", "검토 사유", "조문별 다이렉트 링크"]
# 👆👆👆 여기까지 👆👆👆

HEADERS = {'User-Agent': 'Mozilla/5.0'}
session = requests.Session()
retry = Retry(connect=3, backoff_factor=0.5)
adapter = HTTPAdapter(max_retries=retry)
session.mount('http://', adapter)
session.mount('https://', adapter)

# ==========================================
# 3. 공단 전용 491개 자격 종목 사전
# ==========================================
QNET_CERTS = """
[건설] 금속재창호기능사, 플라스틱창호기능사, 건축구조기술사, 건축기계설비기술사, 건축시공기술사, 건축품질시험기술사, 교통기술사, 농어업토목기술사, 도로및공항기술사, 도시계획기술사, 상하수도기술사, 수자원개발기술사, 조경기술사, 지적기술사, 지질및지반기술사, 철도기술사, 측량및지형공간정보기술사, 토목구조기술사, 토목시공기술사, 토목품질시험기술사, 토질및기초기술사, 항만및해안기술사, 해양기술사, 건축목재시공기능장, 건축일반시공기능장, 배관기능장, 잠수기능장, 건설재료시험기사, 건축기사, 건축설비기사, 교통기사, 도시계획기사, 실내건축기사, 응용지질기사, 조경기사, 지적기사, 철도토목기사, 측량및지형공간정보기사, 콘크리트기사, 토목기사, 항로표지기사, 해양공학기사, 해양자원개발기사, 해양환경기사, 건설재료시험산업기사, 건축목공산업기사, 건축산업기사, 건축설비산업기사, 건축일반시공산업기사, 공간정보융합산업기사, 교통산업기사, 방수산업기사, 배관산업기사, 실내건축산업기사, 잠수산업기사, 조경산업기사, 지적산업기사, 측량및지형공간정보산업기사, 콘크리트산업기사, 토목산업기사, 항로표지산업기사, 해양조사산업기사, 거푸집기능사, 건설재료시험기능사, 건축도장기능사, 건축목공기능사, 공간정보융합기능사, 굴착기운전기능사, 기중기운전기능사, 도배기능사, 도화기능사, 로더운전기능사, 롤러운전기능사, 미장기능사, 방수기능사, 배관기능사, 불도저운전기능사, 비계기능사, 석공기능사, 실내건축기능사, 양화장치운전기능사, 온수온돌기능사, 유리시공기능사, 잠수기능사, 전산응용건축제도기능사, 전산응용토목제도기능사, 조경기능사, 조적기능사, 지게차운전기능사, 지도제작기능사, 지적기능사, 천공기운전기능사, 천장크레인운전기능사, 철근기능사, 철도토목기능사, 측량기능사, 컨테이너크레인운전기능사, 콘크리트기능사, 타워크레인운전기능사, 타일기능사, 항공사진기능사, 항로표지기능사
[경영·회계·사무] 공장관리기술사, 포장기술사, 품질관리기술사, 포장기사, 품질경영기사, 포장산업기사, 품질경영산업기사, 사회조사분석사1급, 사회조사분석사2급, 소비자전문상담사1급, 소비자전문상담사2급, 컨벤션기획사1급, 컨벤션기획사2급
[광업자원] 화약류관리기술사, 화약류관리기사, 화약류관리산업기사, 화약취급기능사
[교육·자연과학·사회과학] 이러닝운영관리사
[기계] 건설기계기술사, 공조냉동기계기술사, 금형기술사, 기계기술사, 산업기계설비기술사, 조선기술사, 차량기술사, 철도차량기술사, 항공기관기술사, 항공기체기술사, 건설기계정비기능장, 금형기능장, 기계가공기능장, 자동차정비기능장, 철도차량정비기능장, 건설기계설비기사, 건설기계정비기사, 공조냉동기계기사, 궤도장비정비기사, 그린전동자동차기사, 농업기계기사, 사출금형기사, 설비보전기사, 승강기기사, 일반기계기사, 자동차정비기사, 조선선체기사, 조선의장기사, 철도차량기사, 프레스금형기사, 항공기사, 건설기계설비산업기사, 건설기계정비산업기사, 공조냉동기계산업기사, 궤도장비정비산업기사, 기계설계산업기사, 기계조립산업기사, 농업기계산업기사, 사출금형산업기사, 설비보전산업기사, 스마트공장산업기사, 승강기산업기사, 자동차정비산업기사, 자동화설비산업기사, 정밀측정산업기사, 조선산업기사, 철도차량산업기사, 컴퓨터응용가공산업기사, 프레스금형산업기사, 항공산업기사, 건설기계정비기능사, 공조냉동기계기능사, 궤도장비정비기능사, 금형기능사, 기계가공조립기능사, 농업기계정비기능사, 반도체설비보전기능사, 선박기관정비기능사, 선체건조기능사, 선체설계기능사, 설비보전기능사, 스마트공장기능사, 승강기기능사, 이륜자동차정비기능사, 자동차보수도장기능사, 자동차정비기능사, 자동차차체수리기능사, 자동화설비기능사, 전산응용기계제도기능사, 정밀측정기능사, 철도차량정비기능사, 컴퓨터응용밀링기능사, 컴퓨터응용선반기능사, 표면실장장비기능사, 항공기정비기능사, 항공전기·전자정비기능사
[농림어업] 농화학기술사, 산림기술사, 수산양식기술사, 시설원예기술사, 어업기술사, 종자기술사, 축산기술사, 산림기능장, 산림기사, 수산양식기사, 시설원예기사, 식물보호기사, 어업생산관리기사, 유기농업기사, 임산가공기사, 임업종묘기사, 종자기사, 축산기사, 화훼장식기사, 버섯산업기사, 산림산업기사, 수산양식산업기사, 식물보호산업기사, 어로산업기사, 유기농업산업기사, 종자산업기사, 축산산업기사, 화훼장식산업기사, 목재가공기능사, 버섯종균기능사, 산림기능사, 수산양식기능사, 식육처리기능사, 원예기능사, 유기농업기능사, 임업종묘기능사, 종자기능사, 축산기능사, 펄프종이제조기능사, 화훼장식기능사
[문화·예술·디자인·방송] 제품디자인기술사, 시각디자인기사, 제품디자인기사, 컬러리스트기사, 시각디자인산업기사, 제품디자인산업기사, 컬러리스트산업기사, 웹디자인개발기능사, 제품응용모델링기능사, 컴퓨터그래픽기능사
[보건·의료] 국제의료관광코디네이터, 임상심리사1급, 임상심리사2급
[사업관리] 공공조달관리사
[사회복지·종교] 직업상담사1급, 직업상담사2급
[섬유·의복] 섬유기술사, 의류기술사, 한복기능장, 섬유기사, 의류기사, 섬유디자인산업기사, 섬유산업기사, 신발산업기사, 패션디자인산업기사, 패션머천다이징산업기사, 남성복기능사, 봉제기능사, 세탁기능사, 신발제조기능사, 여성복기능사, 염색기능사(날염), 염색기능사(침염), 한복기능사
[식품가공] 수산제조기술사, 식품기술사, 제과기능장, 수산제조기사, 식육가공기사, 식품안전기사, 식품산업기사, 제과산업기사, 제빵산업기사, 떡제조기능사, 식품가공기능사, 제과기능사, 제빵기능사
[안전관리] 가스기술사, 건설안전기술사, 기계안전기술사, 비파괴검사기술사, 산업위생관리기술사, 소방기술사, 인간공학기술사, 전기안전기술사, 화공안전기술사, 가스기능장, 가스기사, 건설안전기사, 농작업안전보건기사, 누설비파괴검사기사, 방사선비파괴검사기사, 방재기사, 산업안전기사, 산업위생관리기사, 소방설비기사(기계분야), 소방설비기사(전기분야), 와전류비파괴검사기사, 인간공학기사, 자기비파괴검사기사, 초음파비파괴검사기사, 침투비파괴검사기사, 화재감식평가기사, 가스산업기사, 건설안전산업기사, 방사선비파괴검사산업기사, 산업안전산업기사, 산업위생관리산업기사, 소방설비산업기사(기계분야), 소방설비산업기사(전기분야), 자기비파괴검사산업기사, 초음파비파괴검사산업기사, 침투비파괴검사산업기사, 화재감식평가산업기사, 가스기능사, 방사선비파괴검사기능사, 자기비파괴검사기능사, 초음파비파괴검사기능사, 침투비파괴검사기능사
[영업·판매] 텔레마케팅관리사
[운전·운송] 철도운송산업기사, 농기계운전기능사
[음식서비스] 조리기능장, 복어조리산업기사, 양식조리산업기사, 일식조리산업기사, 중식조리산업기사, 한식조리산업기사, 복어조리기능사, 양식조리기능사, 일식조리기능사, 조주기능사, 중식조리기능사, 한식조리기능사
[이용·숙박·여행·오락·스포츠] 미용장, 이용장, 스포츠경영관리사, 미용사(네일), 미용사(메이크업), 미용사(일반), 미용사(피부), 이용사
[인쇄·목재·가구·공예] 귀금속가공기능장, 인쇄설계기사, 가구제작산업기사, 귀금속가공산업기사, 디지털인쇄산업기사, 보석감정산업기사, 보석디자인산업기사, 피아노조율산업기사, 가구제작기능사, 귀금속가공기능사, 도자공예기능사, 목공예기능사, 보석가공기능사, 보석감정사, 사진기능사, 석공예기능사, 인쇄기능사, 전자출판기능사, 피아노조율기능사
[재료] 금속가공기술사, 금속재료기술사, 금속제련기술사, 세라믹기술사, 용접기술사, 표면처리기술사, 금속재료기능장, 압연기능장, 용접기능장, 제강기능장, 제선기능장, 주조기능장, 판금제관기능장, 표면처리기능장, 금속재료기사, 용접기사, 금속재료산업기사, 용접산업기사, 주조산업기사, 판금제관산업기사, 표면처리산업기사, 가스텅스텐아크용접기능사, 금속도장기능사, 금속재료시험기능사, 압연기능사, 열처리기능사, 이산화탄소가스아크용접기능사, 제강기능사, 제선기능사, 주조기능사, 축로기능사, 판금제관기능사, 표면처리기능사, 피복아크용접기능사
[전기·전자] 건축전기설비기술사, 발송배전기술사, 산업계측제어기술사, 전기응용기술사, 전기철도기술사, 전자응용기술사, 철도신호기술사, 전기기능장, 전자기능장, 광학기사, 로봇기구개발기사, 로봇소프트웨어개발기사, 로봇하드웨어개발기사, 의공기사, 임베디드기사, 전기공사기사, 전기기사, 전기철도기사, 전자기사, 철도신호기사, 3D프린터개발산업기사, 광학기기산업기사, 반도체커스텀레이아웃산업기사, 의공산업기사, 전기공사산업기사, 전기산업기사, 전기철도산업기사, 전자산업기사, 철도신호산업기사, 3D프린터운용기능사, 의료전자기능사, 임베디드기능사, 전기기능사, 전자기능사, 전자캐드기능사, 철도전기신호기능사
[정보통신] 정보관리기술사, 컴퓨터시스템응용기술사, 정보처리기사, 컴퓨터시스템기사, 사무자동화산업기사, 정보처리산업기사, 멀티미디어콘텐츠제작전문가, 정보기기운용기능사, 프로그래밍기능사
[화학] 화공기술사, 위험물기능장, 바이오화학제품제조기사, 정밀화학기사, 화공기사, 화약류제조기사, 화학분석기사, 바이오화학제품제조산업기사, 위험물산업기사, 화약류제조산업기사, 바이오공정기능사, 위험물기능사, 화학분석기능사
[환경·에너지] 기상예보기술사, 대기관리기술사, 소음진동기술사, 수질관리기술사, 자연환경관리기술사, 토양환경기술사, 폐기물처리기술사, 에너지관리기능장, 기상감정기사, 기상기사, 대기환경기사, 생물분류기사(동물), 생물분류기사(식물), 소음진동기사, 수질환경기사, 신재생에너지발전설비기사(태양광), 에너지관리기사, 온실가스관리기사, 자연생태복원기사, 토양환경기사, 폐기물처리기사, 환경위해관리기사, 대기환경산업기사, 소음진동산업기사, 수질환경산업기사, 신재생에너지발전설비산업기사(태양광), 에너지관리산업기사, 자연생태복원산업기사, 폐기물처리산업기사, 신재생에너지발전설비기능사(태양광), 에너지관리기능사, 환경기능사
"""

def get_base_laws():
    all_laws_dict = {}
    print(f"\n📅 [V27.1] {SEARCH_DATE_RANGE} 데이터 수집 시작...")
    
    for target_type in ['law', 'histlaw']:
        page = 1
        while True:
            search_url = f"https://www.law.go.kr/DRF/lawSearch.do?OC={LAW_API_KEY}&target={target_type}&type=XML&efYd={SEARCH_DATE_RANGE}&display=100&page={page}"
            try:
                response = session.get(search_url, headers=HEADERS, timeout=15)
                if not response.text.strip() or response.status_code != 200: break
                root = ET.fromstring(response.text)
                law_nodes = root.findall('.//law')
                if not law_nodes: break
                
                for law in law_nodes:
                    law_id = law.find('법령일련번호').text if law.find('법령일련번호') is not None else ""
                    law_name = law.find('법령명한글').text if law.find('법령명한글') is not None else "이름없음"
                    enforce_date = law.find('시행일자').text if law.find('시행일자') is not None else ""
                    if not law_id or law_name in all_laws_dict: continue
                    
                    law_link = f"https://www.law.go.kr/LSW/lsInfoP.do?lsiSeq={law_id}"
                    detail_url = f"https://www.law.go.kr/DRF/lawService.do?OC={LAW_API_KEY}&target={target_type}&MST={law_id}&type=XML"
                    detail_response = session.get(detail_url, headers=HEADERS, timeout=15)
                    detail_root = ET.fromstring(detail_response.text)
                    
                    reason_text = ""
                    for tag in ['.//개정이유', './/제개정이유']:
                        r_node = detail_root.find(tag)
                        if r_node is not None and r_node.text: reason_text += r_node.text.strip() + "\n"
                    
                    body_text = "\n".join([j.text.strip() for j in detail_root.findall('.//조문내용') if j.text])
                    stars = "\n".join([s.text.strip() for s in detail_root.findall('.//별표내용') if s.text])
                    full_text = f"[개정이유]\n{reason_text}\n[조문내용]\n{body_text}\n[별표]\n{stars}"[:20000]
                    
                    all_laws_dict[law_name] = {"법령명": law_name, "시행일자": enforce_date, "원본": full_text, "링크": law_link}
                    time.sleep(0.1) 
                if len(law_nodes) < 100: break
                page += 1
            except Exception: break
    return list(all_laws_dict.values())

def run_ai_analysis(law, attempt_count=5):
    # 🔥 V26.4 선생님의 오리지널 프롬프트 100% 복원
    prompt = f"""
    당신은 한국산업인력공단의 국가기술자격 규제 심사 연구원입니다.
    아래 제공된 [법령명]과 [내용]을 심층 분석하여, [491개 자격 사전]에 명시된 국가기술자격증과의 연관성을 평가하고 정해진 JSON 양식으로만 출력하십시오.
    
    [491개 자격 사전] {QNET_CERTS}
    
    [법령명] {law['법령명']}
    [내용] {law['원본']}

    [판정 가이드라인]
    1. 분류 기준 (아래 3가지 중 택1):
       - '연관높음': 자격증 소지자의 선임 의무, 가점, 면허/등록 기준 등 활용도가 직접적으로 변동되는 경우
       - '단순관련': 법령에 명칭은 언급되나 실제 활용도 변화가 없는 경우
       - '일반': 자격증과 무관한 경우
    2. 활용도_구분: '연관높음'인 경우에만 [대폭 증가, 소폭 증가, 소폭 감소, 대폭 감소] 중 선택.
    3. 소관부처: 법령 내용을 바탕으로 해당 법령을 소관하는 정부 부처명 추출 (예: 고용노동부, 국토교통부 등)
    4. 개정유형: 법령의 제·개정 성격 추출 (예: 일부개정, 전부개정, 제정, 타법개정 등)
    5. 근거조문: 판단의 결정적인 근거가 된 조항 번호 (예: 제3조제1항, 별표2 등)
    6. 조문번호_숫자: 연관된 조문이 여러 개일 경우 **반드시 모두 추출**하여 배열 형태로 작성 (예: [{{"조문명": "제23조", "숫자": "23"}}, {{"조문명": "제38조의2", "숫자": "38.2"}}])
    7. AI_신뢰도: 본 분석에 대한 AI의 객관적 확신도 ('높음', '보통', '낮음' 중 택1)
       - 높음: 법령 본문이나 별표에 '국가기술자격 종목 명칭'이 정확히 텍스트로 명시된 경우
       - 보통: 종목 명칭이 직접 명시되지는 않았으나, 직무 내용상 연관성이 매우 높다고 강하게 추론되는 경우
       - 낮음: 법령 내용이 모호하거나, 자격증과의 연관성을 억지로 논리적 비약을 통해 연결해야 하는 경우
       
    8. 검토필요: 실무자의 교차 검증이 반드시 필요한 경우 'O', 아니면 'X'
       - [체크(O) 필수 조건]: ① 'AI_신뢰도'가 '보통' 또는 '낮음'이거나, ② '활용도_구분'이 '대폭 증가/감소'로 파급력이 큰 경우
       
    9. 검토사유: '검토필요'가 'O'인 경우에 한해, 그 이유를 작성 (예: "자격 명칭이 직접 명시되지 않아 실무자 확인 요망", "활용도가 대폭 증가하여 정책적 대응 필요" 등). '검토필요'가 'X'이면 빈칸("").
    
    🔥 [작성 가이드라인: 주요 제·개정내용 (요약)] 🔥
    - 실제 개정된 조항과 객관적인 팩트만 글머리 기호('-')를 사용하여 나열하십시오.

    🔥 [작성 가이드라인: 활용도 분석 상세] 🔥
    - [1000자 이내 제한] 1000자 이내로 간결하고 명확하게 분석하십시오.
    - ① 개정 배경, ② 방향성, ③ 파급효과에 집중하십시오.

    [🚨 JSON 작성 절대 규칙]
    1. 출력은 단 1개의 JSON 객체({{ }})만.
    2. (큰따옴표 전면 금지) 모든 텍스트 내부에 절대 큰따옴표(") 금지. 강조는 작은따옴표(') 사용.
    3. (실제 엔터키 금지) 텍스트 내부 실제 줄바꿈 대신 '\\n' 기호 사용.
    4. (종목 포맷팅) 각 직무분야 시작 시 'O ' 꼭지 사용 및 줄바꿈 기호('\\n') 사용.

    [출력 JSON 형태]
    {{
        "분류": "'연관높음', '단순관련', '일반' 중 택 1",
        "소관부처": "고용노동부",
        "개정유형": "일부개정",
        "요약": "- 제O조: 객관적 팩트\\n- 제O조: 객관적 팩트",
        "종목": "O 직무분야: 종목A, 종목B\\nO 직무분야2: 종목C", 
        "활용도_구분": "선택",
        "활용도_분석": "① 개정 배경: ... \\n② 방향성: ... \\n③ 파급효과: ...",
        "근거조문": "제O조 제O항",
        "조문리스트": [
            {{"조문명": "제23조의3", "숫자": "23.3"}},
            {{"조문명": "별표 5", "숫자": ""}}
        ],
        "AI_신뢰도": "보통",
        "검토필요": "O",
        "검토사유": "법령에는 '관련 기술자격'으로만 포괄 명시되어 있어, 정확한 해당 종목 매칭에 대한 실무자 교차 검증이 필요함."
    }}
    """
    
    for attempt in range(attempt_count):
        try:
            response = client.models.generate_content(
                model='gemini-2.5-flash',
                contents=prompt,
                config=types.GenerateContentConfig(response_mime_type="application/json", max_output_tokens=8192)
            )
            raw_text = response.text.strip().replace("```json", "").replace("```", "")
            data = json.loads(raw_text, strict=False)
            
            # 🔥 V26.4 오리지널 다중 링크 파싱 로직 원복
            jomun_list = data.get("조문리스트", [])
            if not jomun_list or not isinstance(jomun_list, list):
                jomun_list = [{"조문명": "내용 확인", "숫자": ""}]
                
            links_str_list = []
            names_str_list = []
            
            for j in jomun_list:
                j_name = j.get("조문명", "확인불가")
                
                # 🔥 [추가된 2줄] AI가 "별표1", "별표   1" 등으로 줘도 무조건 "별표 1"로 교정
                if "별표" in j_name:
                    j_name = re.sub(r'별표\s*(\d+)', r'별표 \1', j_name)
                    
                j_num = str(j.get("숫자", "")).strip().replace(".", ":")
                anchor = f"#J{j_num}" if j_num else ""
                
                # 🔥 [추가 보정] 어색한 '내용 확인' 텍스트를 깔끔하게 다듬기
                if j_name == "내용 확인":
                    names_str_list.append("전체 (세부 조문 미지정)")
                    links_str_list.append(f"▶ {law['법령명']}\n{law['링크']}")
                else:
                    names_str_list.append(j_name)
                    links_str_list.append(f"▶ {law['법령명']} {j_name}\n{law['링크']}{anchor}")
                
            links_str = "\n\n".join(links_str_list)
            names_str = ", ".join(names_str_list)
            
            law_info = {
                "시행일자": law["시행일자"],
                "소관부처": data.get("소관부처", ""),
                "법령명": law["법령명"],
                "개정유형": data.get("개정유형", ""),
                "주요 제·개정내용": data.get("요약", ""),
                "법령 관련 국가기술자격 종목": data.get("종목", ""),
                "활용도 분석 구분": data.get("활용도_구분", ""),
                "활용도 분석 상세": data.get("활용도_분석", ""),
                "근거 조문": names_str,
                "AI 신뢰도": data.get("AI_신뢰도", ""),
                "검토 필요": data.get("검토필요", "X"),
                "조문별 다이렉트 링크": links_str
            }
            return True, data.get("분류", ""), law_info
            
        except Exception as e:
            time.sleep(15)
            error_msg = str(e)
            
    return False, "", {"error": error_msg if 'error_msg' in locals() else "재시도 초과"}

def format_google_sheet(sheet_id, worksheet_name):
    """🔥 구글 시트에 직접 접속하여 헤더 및 정렬 서식 적용"""
    try:
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        creds = ServiceAccountCredentials.from_json_keyfile_dict(json.loads(GCP_SA_JSON), scope)
        gc = gspread.authorize(creds)
        sh = gc.open_by_key(sheet_id)
        ws = sh.worksheet(worksheet_name)
        
        # 헤더 서식 (연한 파랑 배경, 굵게, 가운데 정렬)
        header_fmt = {
            "backgroundColor": {"red": 0.85, "green": 0.88, "blue": 0.95},
            "textFormat": {"bold": True},
            "horizontalAlignment": "CENTER", "verticalAlignment": "MIDDLE"
        }
        ws.format("A1:L1", header_fmt)
        
        # 전체 데이터 정렬 및 줄바꿈
        body_fmt = {
            "horizontalAlignment": "CENTER", "verticalAlignment": "MIDDLE", "wrapStrategy": "WRAP"
        }
        ws.format("A:L", body_fmt)
        print(f"  ✨ [{worksheet_name}] 구글 시트 서식 적용 완료")
    except Exception as e:
        print(f"  ⚠️ 시트 서식 적용 중 오류: {e}")

def write_to_google_sheet(total_len, high_list, simple_list):
    """🔥 구글 시트 마스터 DB 적재 (서식 포함)"""
    if not GCP_SA_JSON or not GOOGLE_SHEET_ID: return
    print("\n📝 구글 시트 마스터 DB 적재 시작...")
    try:
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        creds = ServiceAccountCredentials.from_json_keyfile_dict(json.loads(GCP_SA_JSON), scope)
        client = gspread.authorize(creds)
        sheet = client.open_by_key(GOOGLE_SHEET_ID)

        ws_summary = sheet.worksheet("총괄현황표")
        ws_summary.append_row([FILE_PREFIX, total_len, len(high_list), len(simple_list)])
        
        def prepare_rows(laws):
            return [[law.get(c, "") for c in COLUMNS] for law in laws]

        if high_list:
            ws_high = sheet.worksheet("연관 높은 법령")
            ws_high.append_rows(prepare_rows(high_list))
            format_google_sheet(GOOGLE_SHEET_ID, "연관 높은 법령")
            
        if simple_list:
            ws_simple = sheet.worksheet("국가기술자격 관계 법령(단순 관련)")
            ws_simple.append_rows(prepare_rows(simple_list))
            format_google_sheet(GOOGLE_SHEET_ID, "국가기술자격 관계 법령(단순 관련)")
            
    except Exception as e: print(f"❌ 시트 적재 오류: {e}")

def apply_excel_formatting(filename, total_len, high_list, simple_list):
    """🔥 첨부용 엑셀 파일 서식 완벽 복구 (테두리 추가)"""
    print("\n📊 보고용 엑셀 파일 서식 적용 중...")
    df_summary = pd.DataFrame({"구분": ["총 시행", "연관높음", "단순관련"], "건수": [total_len, len(high_list), len(simple_list)]})
    df_high = pd.DataFrame(high_list) if high_list else pd.DataFrame(columns=COLUMNS)
    df_simple = pd.DataFrame(simple_list) if simple_list else pd.DataFrame(columns=COLUMNS)

    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df_summary.to_excel(writer, sheet_name='총괄현황표', index=False)
        df_high[COLUMNS].to_excel(writer, sheet_name='연관 높은 법령', index=False)
        df_simple[COLUMNS].to_excel(writer, sheet_name='국가기술자격 관계 법령(단순 관련)', index=False)
    
    wb = load_workbook(filename)
    side = Side(style='thin', color="000000")
    border = Border(left=side, right=side, top=side, bottom=side)

    for sheet_name in ['연관 높은 법령', '국가기술자격 관계 법령(단순 관련)']:
        ws = wb[sheet_name]
        for i in range(1, 13):
            cell = ws.cell(row=1, column=i)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.border = border
            
        widths = {'A':12, 'B':15, 'C':35, 'D':12, 'E':45, 'F':40, 'G':18, 'H':50, 'I':18, 'J':12, 'K':10, 'L':55}
        for col, width in widths.items(): ws.column_dimensions[col].width = width
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border
    wb.save(filename)
    return filename

def send_webhook_with_file(fname, total, high, simple):
    """🔥 Make.com에 안전하게 통계와 엑셀 전송 (0건 증발 완벽 방어 에디션)"""
    
    # 숫자 0이 Make.com에서 '빈 값'으로 오해받아 증발하는 것을 막기 위해 
    # 아예 "건" 글자를 강제로 붙여서 '문자열'로 박아버립니다!
    summary_data = {
        "date": str(FILE_PREFIX), 
        "total": f"{total}건", 
        "high": f"{high}건", 
        "simple": f"{simple}건"
    }
    
    print(f"\n🚀 Make.com 전송 시도: 총 {total}건 / 연관 {high}건 / 단순 {simple}건")
    try:
        if fname and os.path.exists(fname):
            with open(fname, 'rb') as f:
                # 엑셀 파일 첨부 전송
                files = {'file': (os.path.basename(fname), f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                response = requests.post(WEBHOOK_URL, data=summary_data, files=files)
        else:
            response = requests.post(WEBHOOK_URL, data=summary_data)
        
        # 🔥 Make.com 서버가 뭐라고 대답했는지 터미널에 강제 출력!
        print(f"  👉 Make.com 서버 응답 코드: {response.status_code}")
        print(f"  👉 Make.com 서버 응답 내용: {response.text}")
        
        if response.status_code == 200:
            print("  ✅ 웹훅 전송 성공! (Make.com 서버가 데이터를 꽉 잡았습니다!)")
        else:
            print(f"  ❌ Make.com 수신 실패 또는 보류됨!")
            
    except Exception as e: 
        print(f"❌ 웹훅 네트워크 에러: {e}")

def main():
    laws = get_base_laws()
    if not laws:
        send_webhook_with_file(None, 0, 0, 0)
        return
    
    high_impact_laws, simple_related_laws, failed_queue = [], [], []
    print(f"\n🏎️  {len(laws)}건 정밀 분석 시작...")
    
    for idx, law in enumerate(laws):
        print(f"[{idx+1}/{len(laws)}] {law['법령명']}... ", end="", flush=True)
        success, cat, law_info = run_ai_analysis(law)
        if success:
            if cat == "연관높음": high_impact_laws.append(law_info); print("🔥")
            elif cat == "단순관련": simple_related_laws.append(law_info); print("🟡")
            else: print("❌")
        else: 
            failed_queue.append(law)
            print(f"⏩ [실패 원인: {law_info.get('error', '알 수 없음')}]")
        time.sleep(5) 
            
    if failed_queue:
        print("\n🚑 패자부활전 시작...")
        time.sleep(20)
        for law in failed_queue:
            success, cat, law_info = run_ai_analysis(law, 3)
            if success:
                if cat == "연관높음": high_impact_laws.append(law_info)
                elif cat == "단순관련": simple_related_laws.append(law_info)
    
    # 1. 시트 적재
    write_to_google_sheet(len(laws), high_impact_laws, simple_related_laws)
    
    # 2. 엑셀 생성
    fname = f"V27.1_법령모니터링_{TARGET_DATE}.xlsx"
    apply_excel_formatting(fname, len(laws), high_impact_laws, simple_related_laws)
    
    # 3. Make.com 전송
    send_webhook_with_file(fname, len(laws), len(high_impact_laws), len(simple_related_laws))
    print(f"\n🎉 [V27.1] 시스템 구동 완료!")

if __name__ == "__main__":
    main()
