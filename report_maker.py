import os
import json
import requests
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

# 💡 1단계 config 파일에서 설정값들을 가져옵니다.
from config import COLUMNS, WEBHOOK_URL, GCP_SA_JSON, GOOGLE_SHEET_ID, TARGET_DATE

# ==========================================
# 1. 구글 시트 마스터 DB 적재 (V28 분류 로직 적용)
# ==========================================
def upload_to_google_sheet(total_len, high_list, simple_list, target_date=TARGET_DATE,
                           status="🟢 정상 작동", log=""):
    """[V29 최종] 3개 시트에 데이터를 분류하여 적재합니다.
    status/log: 총괄현황표에 통신·처리 상태를 함께 기록 (백필 추적용)."""
    if not GCP_SA_JSON or not GOOGLE_SHEET_ID:
        print("  ⚠️ 구글 시트 설정 정보가 없어 적재를 건너뜁니다.")
        return

    try:
        # JSON 파싱 및 인증
        creds_dict = json.loads(GCP_SA_JSON.strip(), strict=False)
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
        client = gspread.authorize(creds)
        spreadsheet = client.open_by_key(GOOGLE_SHEET_ID)

        # 🌟 날짜 가독성 변환 (20260428 -> 2026년_04월_28일)
        display_date = f"{target_date[:4]}년_{target_date[4:6]}월_{target_date[6:]}일"

        # 1) 총괄현황표 기록 [옵션 B] — 같은 날짜 행에 시도 이력 누적
        #    상태 칸 예: "04:13🔴 → 08:47🔴 → 12:31🟢" (한 줄로 그날 이력 전체 확인)
        try:
            from hrdk_law_core.sheets import upsert_daily_summary_row
            # status 문자열에서 심볼만 추출 (🟢/🔴/🟡)
            symbol = "🟢"
            for s in ("🔴", "🟡", "🟢"):
                if s in status:
                    symbol = s
                    break
            upsert_daily_summary_row(
                spreadsheet,
                sheet_name="총괄현황표",
                target_date_display=display_date,
                cols_before_status=[display_date, total_len, len(high_list), len(simple_list)],
                status_symbol=symbol,
                log=log,
            )
        except Exception as se:
            print(f"  ⚠️ '총괄현황표' 시트 기록 실패: {se}")

        # 데이터 변환 함수
        def prepare_rows(laws):
            return [[law.get(col, "") for col in COLUMNS] for law in laws]

        # 2) 연관 높은 법령 적재
        if high_list:
            try:
                ws_high = spreadsheet.worksheet("연관 높은 법령")
                ws_high.append_rows(prepare_rows(high_list), value_input_option="USER_ENTERED")
                print(f"  🔥 연관 높은 법령 {len(high_list)}건 적재 완료")
            except: print("  ⚠️ '연관 높은 법령' 시트를 찾을 수 없습니다.")

        # 3) 단순 관련 법령 적재
        if simple_list:
            try:
                ws_simple = spreadsheet.worksheet("국가기술자격 관계 법령(단순 관련)")
                ws_simple.append_rows(prepare_rows(simple_list), value_input_option="USER_ENTERED")
                print(f"  🟡 단순 관련 법령 {len(simple_list)}건 적재 완료")
            except: print("  ⚠️ '국가기술자격 관계 법령(단순 관련)' 시트를 찾을 수 없습니다.")

    except Exception as e:
        print(f"  ❌ 구글 시트 적재 중 치명적 오류: {e}")

# ==========================================
# 2. 엑셀 파일 생성 함수 (기존과 동일)
# ==========================================
def create_excel_report(high_impact_laws, simple_related_laws, target_date=TARGET_DATE):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "연관 높은 법령"
    ws2 = wb.create_sheet(title="국가기술자격 관계 법령(단순 관련)")
    
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for ws, data_list in [(ws1, high_impact_laws), (ws2, simple_related_laws)]:
        ws.append(COLUMNS)
        for row_idx, info in enumerate(data_list, 2):
            ws.append([info.get(c, "") for c in COLUMNS])
        # 기본 열 너비 설정
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20

    excel_filename = f"V29_법령모니터링_{target_date}.xlsx"
    wb.save(excel_filename)
    return excel_filename

# ==========================================
# 3. 메이크닷컴 웹훅 전송 (기존과 동일)
# ==========================================
def send_webhook_with_file(fname, total, high, simple, target_date=TARGET_DATE):
    if not WEBHOOK_URL: return
    # 🌟 [근본 원인 해결!] 메일/웹훅으로 보낼 때도 사람이 읽기 편한 날짜로 변환해서 쏩니다!
    display_date = f"{target_date[:4]}년 {target_date[4:6]}월 {target_date[6:]}일"
    
    # 이제 Make.com은 "20260428"이 아니라 "2026년 04월 28일" 이라는 데이터를 받게 됩니다!
    # 🏷️ system/source: 두 시스템(RADAR/monitor)을 구분하는 식별값 (메일 제목 분기용)
    summary_data = {
        "system": "law-monitor",
        "source": "monitor",
        "subject": f"[law-monitor] {display_date} 개정법령 활용도 모니터링 (연관 {high}건)",
        "date": display_date, "total": f"{total}건", "high": f"{high}건", "simple": f"{simple}건"
    }
    try:
        if fname and os.path.exists(fname):
            with open(fname, 'rb') as f:
                requests.post(WEBHOOK_URL, data=summary_data, files={'file': (os.path.basename(fname), f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')})
        else:
            requests.post(WEBHOOK_URL, data=summary_data)
        print("  ✅ 웹훅 전송 성공!")
    except Exception as e: print(f"  ❌ 웹훅 에러: {e}")


# ==========================================
# 자격명칭최신화 (담당자가 편집 → 배치가 대장의 종목명을 교체)
# ==========================================
# '자격 명칭 변경' 작업 지시서. 자격증 명칭이 바뀌면 담당자가 여기 적고,
# 배치가 변경시점이 지난 미적용 행을 읽어 대장의 구 종목명을 신 종목명으로 교체합니다.
#   ※ 폐지/통합이라도 자격 자체는 유효하므로 삭제하지 않습니다. 오직 '명칭 교체'만.
import re as _re
from hrdk_law_core.certs import _normalize_cert as _N

UPDATE_SHEET_NAME = "자격명칭최신화"
UPDATE_HEADERS = ["구명칭", "신명칭", "변경시점", "적용여부", "적용일시", "비고"]
# monitor의 종목 칸 + 대상 탭(관련높음 / 단순관련)
_CERT_COL = "법령 관련 국가기술자격 종목"
_LEDGER_TABS = ["연관 높은 법령", "국가기술자격 관계 법령(단순 관련)"]
_DOT_CERTS = ["항공전기·전자정비기능사"]

_UPDATE_EXAMPLES = [
    ["[예시] 전자계산기조직응용기사", "정보처리기사",   "2020-01-01", "",     "",                 "명칭 완전 변경: 구명칭→신명칭. 변경시점이 지나면 대장에서 교체"],
    ["[예시] 정보기기운용기능사",     "정보처리기능사", "2023-01-01", "",     "",                 "다른 자격과 합쳐지며 명칭이 바뀐 경우도 '구명칭→신명칭'으로 적으면 됨"],
    ["[예시] 미래에바뀔종목",         "새이름종목",     "2099-01-01", "",     "",                 "변경시점이 미래면 그날 전엔 적용 안 함(대기)"],
    ["[예시] 이미적용된예시",         "적용된신명칭",   "2020-01-01", "완료", "2026-01-01 00:00", "적용여부=완료 인 행은 다시 처리하지 않음"],
    ["", "", "", "", "", "═══ 실제 입력은 이 줄 아래부터 작성하세요 (위 [예시] 행들은 지우지 마세요) ═══"],
]


def _sheet():
    creds_dict = json.loads(GCP_SA_JSON.strip(), strict=False)
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
    return gspread.authorize(creds).open_by_key(GOOGLE_SHEET_ID)


def ensure_update_sheet_exists():
    """자격명칭최신화 탭이 없으면 헤더 + 예시 행과 함께 생성 (최초 1회용)."""
    if not GCP_SA_JSON or not GOOGLE_SHEET_ID:
        return
    try:
        ss = _sheet()
        try:
            ss.worksheet(UPDATE_SHEET_NAME)
        except gspread.WorksheetNotFound:
            ws = ss.add_worksheet(title=UPDATE_SHEET_NAME, rows=1000, cols=len(UPDATE_HEADERS))
            ws.append_row(UPDATE_HEADERS)
            ws.append_rows(_UPDATE_EXAMPLES)
            print(f"  ✅ '{UPDATE_SHEET_NAME}' 탭 생성 (헤더 + 예시 {len(_UPDATE_EXAMPLES)-1}행 + 구분줄)")
    except Exception as e:
        print(f"  ⚠️ 자격명칭최신화 탭 확인 실패: {e}")


def read_update_instructions():
    """발효(변경시점 지남)+미적용 명칭변경 지시만 반환: [(row_num, 구명칭, 신명칭)]."""
    if not GCP_SA_JSON or not GOOGLE_SHEET_ID:
        return []
    try:
        ss = _sheet()
        ws = ss.worksheet(UPDATE_SHEET_NAME)
        values = ws.get_all_values()
        if len(values) <= 1:
            return []
        header = values[0]
        idx = {h: i for i, h in enumerate(header)}
        def cell(row, name):
            i = idx.get(name)
            return (row[i].strip() if (i is not None and i < len(row)) else "")
        start = 1
        for r_i, row in enumerate(values[1:], start=1):
            gu = cell(row, "구명칭"); bigo = cell(row, "비고")
            if gu.startswith("[예시]") or "실제 입력은" in bigo or "═══" in bigo:
                start = r_i + 1
        from datetime import datetime, timezone, timedelta
        today_digits = datetime.now(timezone(timedelta(hours=9))).strftime("%Y%m%d")
        out = []
        for r_i in range(start, len(values)):
            row = values[r_i]; row_num = r_i + 1
            gu = cell(row, "구명칭"); sin = cell(row, "신명칭")
            done = cell(row, "적용여부"); when = cell(row, "변경시점")
            if not gu or not sin or gu == sin or done:
                continue
            wd = "".join(ch for ch in when if ch.isdigit())[:8]
            if wd and len(wd) == 8 and wd > today_digits:
                continue
            out.append((row_num, gu, sin))
        return out
    except gspread.WorksheetNotFound:
        return []
    except Exception as e:
        print(f"  ⚠️ 자격명칭최신화 시트 읽기 실패: {e}")
        return []


def mark_update_applied(row_nums):
    """적용 완료 행에 적용여부=완료, 적용일시=현재 표시 (재실행 방지)."""
    if not row_nums or not GCP_SA_JSON or not GOOGLE_SHEET_ID:
        return
    try:
        from datetime import datetime, timezone, timedelta
        ss = _sheet(); ws = ss.worksheet(UPDATE_SHEET_NAME)
        header = ws.row_values(1)
        col_done = header.index("적용여부") + 1
        col_when = header.index("적용일시") + 1
        now = datetime.now(timezone(timedelta(hours=9))).strftime("%Y-%m-%d %H:%M")
        body = []
        for rn in row_nums:
            body.append({"range": gspread.utils.rowcol_to_a1(rn, col_done), "values": [["완료"]]})
            body.append({"range": gspread.utils.rowcol_to_a1(rn, col_when), "values": [[now]]})
        ws.batch_update(body, value_input_option="USER_ENTERED")
    except Exception as e:
        print(f"  ⚠️ 적용여부 기록 실패: {e}")


def _split_cert_cell(raw):
    s = str(raw or "")
    for dc in _DOT_CERTS:
        s = s.replace(dc, dc.replace("·", "㉿"))
    s = _re.sub(r"\(([^)]*)\)", lambda m: "(" + m.group(1).replace(",", "§") + ")", s)
    return [c.strip().replace("§", ",").replace("㉿", "·") for c in _re.split(r"[,/·\n]", s) if c.strip()]


def _apply_cell(cell_value, rename_map):
    """명칭 교체(+칸 내 중복 제거). 삭제는 하지 않음. 반환:(새 값, 변경여부)"""
    items = _split_cert_cell(cell_value)
    if not items:
        return cell_value, False
    out, seen, changed = [], set(), False
    for it in items:
        key = _N(it)
        if key in rename_map:
            new = rename_map[key]
            if _N(new) != key:
                changed = True
            it = new
        nk = _N(it)
        if nk in seen:
            changed = True
            continue
        seen.add(nk); out.append(it)
    new_cell = ", ".join(out)
    if new_cell != str(cell_value or "").strip():
        changed = True
    return new_cell, changed


def run_name_updates():
    """
    자격명칭최신화 탭의 발효·미적용 명칭변경을 대장 2개 탭(연관높음/단순관련)에 반영.
      · 명칭 교체만(구→신), 삭제 없음, 교체 후 칸 내 중복 제거
      · 처리한 지시는 적용여부=완료 표시(1회성)
    환경변수 NAME_UPDATE_PREVIEW=1 이면 미리보기만(시트 수정 안 함).
    """
    if not GCP_SA_JSON or not GOOGLE_SHEET_ID:
        return
    instrs = read_update_instructions()
    if not instrs:
        return
    preview = os.environ.get("NAME_UPDATE_PREVIEW", "").strip() in ("1", "true", "True")
    rename_map, row_nums = {}, []
    for row_num, gu, sin in instrs:
        row_nums.append(row_num); rename_map[_N(gu)] = sin
    print(f"  🔤 자격명칭최신화: 적용할 명칭변경 {len(instrs)}건" + ("  [미리보기]" if preview else ""))

    ss = _sheet()
    total_changed = 0
    for tab in _LEDGER_TABS:
        try:
            ws = ss.worksheet(tab)
        except Exception:
            print(f"    ⚠️ '{tab}' 탭 없음 → 건너뜀")
            continue
        values = ws.get_all_values()
        if len(values) <= 1:
            continue
        header = values[0]
        if _CERT_COL not in header:
            print(f"    ⚠️ '{tab}'에 '{_CERT_COL}' 칸 없음 → 건너뜀")
            continue
        ci = header.index(_CERT_COL)
        col_letter = gspread.utils.rowcol_to_a1(1, ci + 1).rstrip("1")
        changes, plist = [], []
        for r_i in range(1, len(values)):
            row = values[r_i]
            old = row[ci] if ci < len(row) else ""
            new, ch = _apply_cell(old, rename_map)
            if ch and new != (old or "").strip():
                changes.append((r_i + 1, new)); plist.append((r_i + 1, old, new))
        print(f"    • '{tab}': {len(changes)}개 행 변경" + (" (미리보기)" if preview else ""))
        for rn, old, new in plist[:5]:
            print(f"        행{rn}: {old[:30]} → {new[:30]}")
        if not preview and changes:
            body = [{"range": f"{col_letter}{rn}", "values": [[nv]]} for rn, nv in changes]
            for i in range(0, len(body), 500):
                ws.batch_update(body[i:i+500], value_input_option="USER_ENTERED")
        total_changed += len(changes)

    if not preview:
        mark_update_applied(row_nums)
        print(f"    • 지시 {len(row_nums)}건 '완료' 표시 (총 {total_changed}행 변경)")


def read_all_aliases_for_resolve():
    """
    자격명칭최신화 탭에서 '분석 변환용' 별칭 전부({구명칭: 신명칭}).
      · 적용여부=완료 인 것도 포함(과거 이관분도 변환엔 계속 사용)
      · 변경시점이 미래인 것만 제외
      · 예시행/구분줄 위쪽 건너뜀
    """
    if not GCP_SA_JSON or not GOOGLE_SHEET_ID:
        return {}
    try:
        ss = _sheet()
        ws = ss.worksheet(UPDATE_SHEET_NAME)
        values = ws.get_all_values()
        if len(values) <= 1:
            return {}
        header = values[0]
        idx = {h: i for i, h in enumerate(header)}
        def cell(row, name):
            i = idx.get(name)
            return (row[i].strip() if (i is not None and i < len(row)) else "")
        start = 1
        for r_i, row in enumerate(values[1:], start=1):
            gu = cell(row, "구명칭"); bigo = cell(row, "비고")
            if gu.startswith("[예시]") or "실제 입력은" in bigo or "═══" in bigo:
                start = r_i + 1
        from datetime import datetime, timezone, timedelta
        today_digits = datetime.now(timezone(timedelta(hours=9))).strftime("%Y%m%d")
        out = {}
        for r_i in range(start, len(values)):
            row = values[r_i]
            gu = cell(row, "구명칭"); sin = cell(row, "신명칭"); when = cell(row, "변경시점")
            if not gu or not sin or gu == sin:
                continue
            wd = "".join(ch for ch in when if ch.isdigit())[:8]
            if wd and len(wd) == 8 and wd > today_digits:
                continue
            out[gu] = sin
        return out
    except gspread.WorksheetNotFound:
        return {}
    except Exception as e:
        print(f"  ⚠️ 자격명칭최신화(변환용) 읽기 실패: {e}")
        return {}
